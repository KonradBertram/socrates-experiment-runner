"""Core logic for the Socrates Experiment Runner.

The functions in this module are deliberately independent from Streamlit so they
can be unit-tested and reused by the UI layer.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import random
import re
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Any, Iterable, Mapping, Sequence
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_NAME = "Socrates Experiment Runner"
MODEL_ID = "socratesft/socrates-qwen2.5-14b-sft"
API_BASE = "https://api.featherless.ai/v1"

SYSTEM_PROMPT = (
    "You are simulating a survey respondent. Answer exactly as instructed, "
    "following the specified response format without additional commentary."
)

DEFAULT_TEMPERATURE = 0.6
DEFAULT_TOP_P = 0.9
DEFAULT_USE_TOP_K = False
DEFAULT_TOP_K = 50
MAX_NEW_TOKENS = 8
ESTIMATED_COMPLETION_TOKENS_PER_RESPONSE = 2

MAX_HTTP_RETRIES = 4
MAX_APP_WORKERS = 16
MAX_SIMULATIONS = 1_000
MAX_SLOT_ATTEMPTS = 3
MAX_RESPONSE_OPTIONS = 10
OPTION_CODES = tuple(str(i) for i in range(1, 11))

US_PRESET_NOTE = (
    "Approximate, rounded U.S. adult/household marginal distributions. "
    "Each demographic dimension is sampled independently, so the preset does "
    "not preserve real-world correlations between age, income, employment, "
    "education, household size, and housing tenure. It is not the empirical "
    "joint distribution of SocSci210 respondents."
)

US_PRESET_SOURCES = (
    "https://data.census.gov/table/ACSST1Y2023.S0101",
    "https://data.census.gov/table/ACSST1Y2023.S1501",
    "https://data.census.gov/table/ACSST1Y2023.S1901",
    "https://data.census.gov/table/ACSST1Y2023.S1101",
    "https://data.census.gov/table/ACSST1Y2023.B25003",
    "https://www.bls.gov/cps/",
)


# Category labels intentionally follow or closely map to the labels reported in
# the Socrates paper/model prompt. The U.S. preset is a rounded approximation,
# not a reproduction of the training sample.
DEMOGRAPHIC_DEFINITIONS: dict[str, dict[str, Any]] = {
    "Age": {
        "prompt_label": "Age",
        "categories": ["18-24", "25-34", "35-49", "50-64", "65+"],
        "preset": [12.0, 18.0, 25.0, 24.0, 21.0],
    },
    "Gender": {
        "prompt_label": "Gender",
        "categories": ["Female", "Male"],
        "preset": [51.0, 49.0],
    },
    "Income": {
        "prompt_label": "Income",
        "categories": [
            "<5K",
            "5-9K",
            "10-14K",
            "15-19K",
            "20-29K",
            "30-39K",
            "40-49K",
            "50-74K",
            "75-99K",
            "100-124K",
            "125-149K",
            "150-175K+",
            "175-200K+",
            "200K+",
        ],
        "preset": [2.0, 2.0, 3.0, 3.0, 8.0, 7.0, 6.0, 16.0, 13.0, 10.0, 8.0, 6.0, 4.0, 12.0],
    },
    "Education": {
        "prompt_label": "Education",
        "categories": [
            "Less than high school",
            "Some high school (no diploma)",
            "High school graduate or equivalent",
            "Vocational/tech school/some college/associates",
            "Bachelor's degree",
            "Post grad study/professional degree",
        ],
        "preset": [4.0, 7.0, 27.0, 28.0, 22.0, 12.0],
    },
    "Employment": {
        "prompt_label": "Employment",
        "categories": [
            "Employed as paid employee",
            "Self-employed",
            "Looking for work",
            "Temporarily laid off",
            "Retired",
            "Disabled",
            "Not working for other reasons",
        ],
        "preset": [57.0, 6.0, 3.0, 1.0, 18.0, 5.0, 10.0],
    },
    "Marital Status": {
        "prompt_label": "Marital Status",
        "categories": [
            "Married",
            "Never married",
            "Living with partner",
            "Divorced",
            "Widowed",
            "Separated",
        ],
        "preset": [46.0, 28.0, 9.0, 10.0, 5.0, 2.0],
    },
    "Household Size": {
        "prompt_label": "Household Size",
        "categories": ["1", "2", "3", "4", "5+"],
        "preset": [29.0, 34.0, 15.0, 13.0, 9.0],
    },
    "Housing Ownership": {
        "prompt_label": "Housing Ownership",
        "categories": [
            "Owned or being bought by you/someone in your household",
            "Rented for cash",
            "Occupied without payment of cash rent",
        ],
        "preset": [66.0, 32.0, 2.0],
    },
}

DEFAULT_DEMOGRAPHIC_DIMENSIONS = tuple(DEMOGRAPHIC_DEFINITIONS.keys())


class ConfigurationError(ValueError):
    """Raised when a run configuration is internally inconsistent."""


def api_headers(api_key: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "X-Title": APP_NAME,
    }


def config_fingerprint(config: Mapping[str, Any]) -> str:
    payload = json.dumps(config, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def heuristic_token_count(text: str) -> int:
    # A conservative, model-agnostic fallback for English text.
    return max(1, math.ceil(len(text or "") / 3.7))


def tokenize_count(api_key: str, text: str) -> int:
    response = requests.post(
        f"{API_BASE}/tokenize",
        headers=api_headers(api_key),
        json={"model": MODEL_ID, "text": text},
        timeout=30,
    )
    response.raise_for_status()
    tokens = response.json().get("tokens")
    if not isinstance(tokens, list):
        raise ValueError("Tokenizer response did not contain a token list.")
    count = len(tokens)
    if text.strip() and count <= 0:
        raise ValueError("Tokenizer returned zero tokens for non-empty text.")
    return max(1, count)


def get_model_and_plan(api_key: str) -> tuple[dict[str, Any], dict[str, Any]]:
    encoded_model = quote(MODEL_ID, safe="")
    headers = api_headers(api_key)
    model_response = requests.get(
        f"{API_BASE}/models/{encoded_model}",
        headers=headers,
        timeout=30,
    )
    model_response.raise_for_status()
    plan_response = requests.get(
        f"{API_BASE}/plan",
        headers=headers,
        timeout=30,
    )
    plan_response.raise_for_status()
    return model_response.json(), plan_response.json()


def get_runtime_info(api_key: str) -> dict[str, Any]:
    try:
        model_info, plan_info = get_model_and_plan(api_key)
        pricing = model_info.get("pricing", {}) or {}
        prompt_price = float(pricing.get("prompt", 0) or 0)
        completion_price = float(pricing.get("completion", 0) or 0)

        plan_concurrency = int(plan_info.get("concurrency", 1) or 1)
        model_concurrency_cost = int(model_info.get("concurrency_cost", 1) or 1)
        available_workers = max(1, plan_concurrency // max(1, model_concurrency_cost))
        workers = min(MAX_APP_WORKERS, available_workers)

        model_context = model_info.get("context_length")
        plan_context = plan_info.get("max_context_length")
        context_candidates = [
            int(value)
            for value in (model_context, plan_context)
            if value is not None and int(value) > 0
        ]
        effective_context = min(context_candidates) if context_candidates else None

        return {
            "model_info": model_info,
            "plan_info": plan_info,
            "prompt_price": prompt_price,
            "completion_price": completion_price,
            "pricing_available": prompt_price > 0 or completion_price > 0,
            "workers": workers,
            "available_on_current_plan": model_info.get("available_on_current_plan", True),
            "effective_context_length": effective_context,
            "error": None,
        }
    except Exception as exc:  # The app can still attempt a run sequentially.
        return {
            "model_info": {},
            "plan_info": {},
            "prompt_price": 0.0,
            "completion_price": 0.0,
            "pricing_available": False,
            "workers": 1,
            "available_on_current_plan": None,
            "effective_context_length": None,
            "error": str(exc),
        }


def parse_answer_options(text: str) -> list[str]:
    options = [line.strip() for line in (text or "").splitlines() if line.strip()]
    return options


def largest_remainder_counts(total: int, percentages: Sequence[float]) -> list[int]:
    if total < 0:
        raise ValueError("total must be non-negative")
    if not percentages:
        return []
    exact = [total * float(p) / 100.0 for p in percentages]
    counts = [math.floor(value) for value in exact]
    remainder = total - sum(counts)
    order = sorted(
        range(len(exact)),
        key=lambda i: (exact[i] - counts[i], -i),
        reverse=True,
    )
    for index in order[:remainder]:
        counts[index] += 1
    return counts


def distributions_from_preset(dimensions: Iterable[str]) -> dict[str, dict[str, float]]:
    distributions: dict[str, dict[str, float]] = {}
    for dimension in dimensions:
        definition = DEMOGRAPHIC_DEFINITIONS[dimension]
        distributions[dimension] = {
            category: float(percentage)
            for category, percentage in zip(definition["categories"], definition["preset"])
        }
    return distributions


def validate_distribution(distribution: Mapping[str, float], tolerance: float = 0.01) -> bool:
    total = sum(float(value) for value in distribution.values())
    return abs(total - 100.0) <= tolerance and all(float(value) >= 0 for value in distribution.values())


def sample_prompt_value(dimension: str, segment: str, rng: random.Random) -> str:
    if dimension == "Age":
        ranges = {
            "18-24": (18, 24),
            "25-34": (25, 34),
            "35-49": (35, 49),
            "50-64": (50, 64),
            "65+": (65, 90),
        }
        lower, upper = ranges[segment]
        return str(rng.randint(lower, upper))
    if dimension == "Household Size":
        if segment == "5+":
            return str(rng.randint(5, 8))
        return segment
    return segment


def generate_profiles(
    count: int,
    dimensions: Sequence[str],
    distributions: Mapping[str, Mapping[str, float]],
    rng: random.Random,
) -> list[dict[str, dict[str, str]]]:
    if count <= 0:
        return []
    profiles = [
        {"segments": {}, "values": {}}
        for _ in range(count)
    ]

    for dimension in dimensions:
        definition = DEMOGRAPHIC_DEFINITIONS[dimension]
        categories = list(definition["categories"])
        distribution = distributions[dimension]
        percentages = [float(distribution.get(category, 0.0)) for category in categories]
        if not validate_distribution(dict(zip(categories, percentages))):
            raise ConfigurationError(f"{dimension} percentages must total 100%.")

        counts = largest_remainder_counts(count, percentages)
        assignments: list[str] = []
        for category, category_count in zip(categories, counts):
            assignments.extend([category] * category_count)
        rng.shuffle(assignments)

        for profile, segment in zip(profiles, assignments):
            profile["segments"][dimension] = segment
            profile["values"][dimension] = sample_prompt_value(dimension, segment, rng)

    return profiles


def counterbalanced_option_orders(
    options: Sequence[str],
    count: int,
    rng: random.Random,
) -> list[list[str]]:
    """Create near-exactly balanced answer positions using cyclic Latin blocks.

    Each full block contains every cyclic rotation of the options and every
    cyclic rotation of the reversed options. Consequently, every option appears
    equally often in every response-code position within a full block.
    """

    if count <= 0:
        return []
    option_list = list(options)
    if len(option_list) < 2:
        raise ConfigurationError("At least two response options are required.")

    def rotations(items: list[str]) -> list[list[str]]:
        return [items[offset:] + items[:offset] for offset in range(len(items))]

    base_block = rotations(option_list) + rotations(list(reversed(option_list)))
    orders: list[list[str]] = []
    while len(orders) < count:
        block = [list(order) for order in base_block]
        rng.shuffle(block)
        orders.extend(block)
    return orders[:count]


def build_user_prompt(
    config: Mapping[str, Any],
    variant: Mapping[str, Any],
    profile_values: Mapping[str, str],
    option_order: Sequence[str],
) -> tuple[str, dict[str, str]]:
    codes = OPTION_CODES[: len(option_order)]
    code_to_option = dict(zip(codes, option_order))

    profile_lines = []
    for dimension in config["demographic_dimensions"]:
        prompt_label = DEMOGRAPHIC_DEFINITIONS[dimension]["prompt_label"]
        profile_lines.append(f"- {prompt_label}: {profile_values[dimension]}")
    profile_text = "\n".join(profile_lines) if profile_lines else "- No demographic profile supplied"

    option_lines = "\n".join(f"{code}: {code_to_option[code]}" for code in codes)
    allowed_codes = ", ".join(codes)

    context_parts = [
        f"Line of business: {config['line_of_business']}",
        f"Customer journey type: {config['customer_journey_type']}",
    ]
    if config.get("context"):
        context_parts.append(config["context"])
    context_text = "\n".join(context_parts)

    user_prompt = f"""You are a survey respondent with the following demographic profile:
{profile_text}

Read the situation below and answer exactly as this person would. Respond based on what this person would most likely do, not what they should do. Follow the response instructions precisely.

SITUATION CONTEXT
{context_text}

TOUCHPOINT / INTERVENTION SHOWN TO YOU
--- BEGIN VARIANT ---
{variant['text'].strip()}
--- END VARIANT ---

QUESTION
{config['outcome_question'].strip()}

RESPONSE OPTIONS
{option_lines}

Only return the integer corresponding to your answer from this list: {allowed_codes}. Return the integer only, nothing else."""

    return user_prompt, code_to_option


def serialize_messages_for_tokenizer(system_prompt: str, user_prompt: str) -> str:
    # The endpoint does not expose chat-template tokenization directly. Role
    # markers make this closer to the actual chat request than plain concatenation.
    return f"[SYSTEM]\n{system_prompt}\n\n[USER]\n{user_prompt}\n\n[ASSISTANT]\n"


def build_simulation_plan(config: Mapping[str, Any], seed: int) -> list[dict[str, Any]]:
    rng = random.Random(seed)
    jobs: list[dict[str, Any]] = []
    slot_number = 1
    respondent_number = 1

    for variant_index, variant in enumerate(config["variants"]):
        count = int(variant["count"])
        variant_rng = random.Random(rng.getrandbits(64))
        profiles = generate_profiles(
            count=count,
            dimensions=config["demographic_dimensions"],
            distributions=config["demographic_distributions"],
            rng=variant_rng,
        )
        option_orders = counterbalanced_option_orders(
            config["answer_options"],
            count,
            variant_rng,
        )

        paired = list(zip(profiles, option_orders))
        variant_rng.shuffle(paired)
        for profile, option_order in paired:
            user_prompt, code_to_option = build_user_prompt(
                config=config,
                variant=variant,
                profile_values=profile["values"],
                option_order=option_order,
            )
            jobs.append(
                {
                    "slot_id": f"S{slot_number:05d}",
                    "respondent_id": f"R{respondent_number:05d}",
                    "slot_attempt": 1,
                    "variant": variant["name"],
                    "variant_index": variant_index,
                    "profile_segments": profile["segments"],
                    "profile_values": profile["values"],
                    "option_order": list(option_order),
                    "code_to_option": code_to_option,
                    "system_prompt": SYSTEM_PROMPT,
                    "user_prompt": user_prompt,
                }
            )
            slot_number += 1
            respondent_number += 1

    rng.shuffle(jobs)
    return jobs


def normalize_model_answer(
    raw_text: str,
    code_to_option: Mapping[str, str],
) -> tuple[str | None, str | None]:
    """Map a Socrates completion to one of the allowed response options.

    Socrates was fine-tuned on direct survey prediction prompts whose answers are
    commonly numeric labels. We therefore prefer exact numeric outputs, but also
    tolerate a small set of harmless wrappers such as "Option 2" or
    "The answer is 2" so formatting noise does not burn simulation slots.
    """
    text = (raw_text or "").strip()
    if not text:
        return None, None

    text = text.replace("```", "").strip()
    first_line = next((line.strip() for line in text.splitlines() if line.strip()), "")
    cleaned = first_line.strip(" \t\r\n\"'`.,;:()[]{}")

    # Best case: the model follows the instruction exactly.
    for code, option in code_to_option.items():
        if cleaned.casefold() == str(code).casefold():
            return str(code), option

    # Also accept the exact option text if the model returns the label rather than
    # the requested code.
    for code, option in code_to_option.items():
        if cleaned.casefold() == option.strip().casefold():
            return str(code), option

    allowed = sorted((str(code) for code in code_to_option), key=len, reverse=True)
    escaped = "|".join(re.escape(code) for code in allowed)

    # Common wrappers produced by instruction-tuned chat models.
    wrapper_patterns = [
        rf"^\s*(?:option|choice|response|answer)\s*[:=#-]?\s*({escaped})(?:\b|$)",
        rf"^\s*(?:i\s+(?:would\s+)?(?:choose|select|pick))\s*[:=#-]?\s*(?:option\s*)?({escaped})(?:\b|$)",
        rf"^\s*(?:the\s+)?(?:answer|choice|response)\s+(?:is|would\s+be)\s*[:=#-]?\s*({escaped})(?:\b|$)",
        rf"^\s*({escaped})\s*[\).,:;-]",
    ]
    for pattern in wrapper_patterns:
        match = re.search(pattern, first_line, flags=re.IGNORECASE)
        if match:
            candidate = match.group(1)
            if candidate in code_to_option:
                return candidate, code_to_option[candidate]

    return None, None


def extract_chat_content(choice: Mapping[str, Any]) -> str:
    """Extract text robustly from OpenAI-compatible chat completion shapes."""
    message = choice.get("message") or {}
    content = message.get("content") if isinstance(message, Mapping) else None

    if isinstance(content, str):
        return content

    # Some OpenAI-compatible providers may return a list of content parts.
    if isinstance(content, list):
        parts: list[str] = []
        for part in content:
            if isinstance(part, str):
                parts.append(part)
            elif isinstance(part, Mapping):
                value = part.get("text") or part.get("content")
                if isinstance(value, str):
                    parts.append(value)
        if parts:
            return "".join(parts)

    # Defensive fallback for providers exposing completion text directly.
    direct_text = choice.get("text")
    return direct_text if isinstance(direct_text, str) else ""


def run_one(
    job: Mapping[str, Any],
    api_key: str,
    settings: Mapping[str, Any],
    runtime_info: Mapping[str, Any],
) -> dict[str, Any]:
    data: dict[str, Any] = {
        "model": MODEL_ID,
        "messages": [
            {"role": "system", "content": job["system_prompt"]},
            {"role": "user", "content": job["user_prompt"]},
        ],
        "max_tokens": MAX_NEW_TOKENS,
        "temperature": float(settings["temperature"]),
        "top_p": float(settings["top_p"]),
    }
    if settings.get("use_top_k"):
        data["top_k"] = int(settings["top_k"])

    retryable_statuses = {408, 425, 429, 500, 502, 503, 504}
    retry_events = 0
    http_attempts = 0
    last_error: str | None = None

    for attempt in range(MAX_HTTP_RETRIES):
        http_attempts += 1
        try:
            response = requests.post(
                f"{API_BASE}/chat/completions",
                headers=api_headers(api_key),
                json=data,
                timeout=(20, 180),
            )

            if response.status_code in retryable_statuses:
                last_error = f"Temporary Featherless error {response.status_code}: {response.text[:300]}"
                if attempt < MAX_HTTP_RETRIES - 1:
                    retry_events += 1
                    time.sleep((1.7**attempt) + random.uniform(0.1, 0.8))
                    continue
                break

            response.raise_for_status()
            payload = response.json()
            choice = payload["choices"][0]
            raw = extract_chat_content(choice)
            selected_code, selected_option = normalize_model_answer(raw, job["code_to_option"])

            usage = payload.get("usage", {}) or {}
            prompt_tokens = int(usage.get("prompt_tokens", 0) or 0)
            completion_tokens = int(usage.get("completion_tokens", 0) or 0)

            cost = None
            if runtime_info.get("pricing_available"):
                cost = (
                    prompt_tokens * float(runtime_info["prompt_price"])
                    + completion_tokens * float(runtime_info["completion_price"])
                )

            status = "valid" if selected_option is not None else "invalid_output"
            return {
                **dict(job),
                "raw_completion": raw,
                "selected_code": selected_code,
                "selected_option": selected_option,
                "status": status,
                "error": None,
                "prompt_tokens": prompt_tokens,
                "completion_tokens": completion_tokens,
                "cost": cost,
                "http_attempts": http_attempts,
                "retry_events": retry_events,
                "finish_reason": choice.get("finish_reason"),
            }

        except requests.RequestException as exc:
            last_error = str(exc)
            status_code = getattr(getattr(exc, "response", None), "status_code", None)
            should_retry = status_code is None or status_code in retryable_statuses
            if should_retry and attempt < MAX_HTTP_RETRIES - 1:
                retry_events += 1
                time.sleep((1.7**attempt) + random.uniform(0.1, 0.8))
                continue
            break
        except (KeyError, IndexError, TypeError, ValueError, json.JSONDecodeError) as exc:
            last_error = f"Unexpected API response: {exc}"
            if attempt < MAX_HTTP_RETRIES - 1:
                retry_events += 1
                time.sleep((1.7**attempt) + random.uniform(0.1, 0.8))
                continue
            break
        except Exception as exc:  # Defensive: keep the batch alive.
            last_error = str(exc)
            if attempt < MAX_HTTP_RETRIES - 1:
                retry_events += 1
                time.sleep((1.7**attempt) + random.uniform(0.1, 0.8))
                continue
            break

    return {
        **dict(job),
        "raw_completion": "",
        "selected_code": None,
        "selected_option": None,
        "status": "api_error",
        "error": last_error or "Unknown API error",
        "prompt_tokens": 0,
        "completion_tokens": 0,
        "cost": None,
        "http_attempts": http_attempts,
        "retry_events": retry_events,
        "finish_reason": None,
    }


def _sample_jobs_by_variant(jobs: Sequence[Mapping[str, Any]], samples_per_variant: int = 3) -> dict[str, list[Mapping[str, Any]]]:
    grouped: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for job in jobs:
        grouped[str(job["variant"])].append(job)

    selected: dict[str, list[Mapping[str, Any]]] = {}
    for variant, variant_jobs in grouped.items():
        ordered = sorted(variant_jobs, key=lambda item: item["slot_id"])
        if len(ordered) <= samples_per_variant:
            selected[variant] = ordered
            continue
        indices = sorted({0, len(ordered) // 2, len(ordered) - 1})
        selected[variant] = [ordered[index] for index in indices[:samples_per_variant]]
    return selected


def estimate_run(
    api_key: str,
    jobs: Sequence[Mapping[str, Any]],
    runtime_info: Mapping[str, Any],
) -> dict[str, Any]:
    samples = _sample_jobs_by_variant(jobs)
    condition_estimates: list[dict[str, Any]] = []
    total_prompt_tokens = 0
    used_fallback = False

    variant_counts = Counter(str(job["variant"]) for job in jobs)
    for variant, sample_jobs in samples.items():
        token_counts: list[int] = []
        sources: list[str] = []
        for job in sample_jobs:
            serialized = serialize_messages_for_tokenizer(job["system_prompt"], job["user_prompt"])
            try:
                count = tokenize_count(api_key, serialized)
                source = "Featherless tokenizer"
            except Exception:
                count = heuristic_token_count(serialized)
                source = "character-based fallback"
                used_fallback = True
            token_counts.append(count)
            sources.append(source)

        average_tokens = max(1, round(sum(token_counts) / len(token_counts)))
        target_count = int(variant_counts[variant])
        estimated_input = average_tokens * target_count
        total_prompt_tokens += estimated_input
        condition_estimates.append(
            {
                "variant": variant,
                "sampled_prompts": len(sample_jobs),
                "estimated_tokens_per_prompt": average_tokens,
                "target_valid_n": target_count,
                "estimated_input_tokens": estimated_input,
                "token_source": (
                    "character-based fallback"
                    if "character-based fallback" in sources
                    else "Featherless tokenizer"
                ),
            }
        )

    total_target = len(jobs)
    estimated_completion_tokens = total_target * ESTIMATED_COMPLETION_TOKENS_PER_RESPONSE
    safety_prompt_tokens = total_prompt_tokens * MAX_SLOT_ATTEMPTS
    safety_completion_tokens = total_target * MAX_SLOT_ATTEMPTS * MAX_NEW_TOKENS

    base_cost = None
    safety_ceiling_cost = None
    if runtime_info.get("pricing_available"):
        base_cost = (
            total_prompt_tokens * float(runtime_info["prompt_price"])
            + estimated_completion_tokens * float(runtime_info["completion_price"])
        )
        safety_ceiling_cost = (
            safety_prompt_tokens * float(runtime_info["prompt_price"])
            + safety_completion_tokens * float(runtime_info["completion_price"])
        )

    return {
        "prompt_tokens": total_prompt_tokens,
        "completion_tokens": estimated_completion_tokens,
        "safety_prompt_tokens": safety_prompt_tokens,
        "safety_completion_tokens": safety_completion_tokens,
        "base_cost": base_cost,
        "safety_ceiling_cost": safety_ceiling_cost,
        "used_fallback": used_fallback,
        "condition_estimates": condition_estimates,
    }


def overall_result_rows(config: Mapping[str, Any], accepted_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    target = config["target_behavior"]
    control = config["control_variant"]
    options = list(config["answer_options"])

    by_variant: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in accepted_rows:
        by_variant[str(row["variant"])].append(row)

    control_rows = by_variant.get(control, [])
    control_n = len(control_rows)
    control_target_count = sum(row.get("selected_option") == target for row in control_rows)
    control_rate = control_target_count / control_n * 100 if control_n else None

    output: list[dict[str, Any]] = []
    for variant in config["variants"]:
        name = variant["name"]
        rows = by_variant.get(name, [])
        valid_n = len(rows)
        counts = Counter(row.get("selected_option") for row in rows)
        target_count = counts.get(target, 0)
        target_rate = target_count / valid_n * 100 if valid_n else None
        effect = None
        if target_rate is not None and control_rate is not None:
            effect = target_rate - control_rate

        result: dict[str, Any] = {
            "Variant": name,
            "Control": name == control,
            "Valid N": valid_n,
            "Target N": int(variant["count"]),
            "Target behavior": target,
            "Target count": target_count,
            "Target rate (%)": target_rate,
            "Effect vs control (pp)": effect,
        }
        for option in options:
            option_count = counts.get(option, 0)
            result[f"{option} count"] = option_count
            result[f"{option} rate (%)"] = option_count / valid_n * 100 if valid_n else None
        output.append(result)
    return output


def segment_result_rows(config: Mapping[str, Any], accepted_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    target = config["target_behavior"]
    control = config["control_variant"]
    variants = [variant["name"] for variant in config["variants"]]
    options = list(config["answer_options"])
    output: list[dict[str, Any]] = []

    for dimension in config["demographic_dimensions"]:
        segment_order = list(config["demographic_distributions"][dimension].keys())
        for segment in segment_order:
            segment_rows = [
                row
                for row in accepted_rows
                if row.get("profile_segments", {}).get(dimension) == segment
            ]
            control_rows = [row for row in segment_rows if row["variant"] == control]
            control_n = len(control_rows)
            control_target_count = sum(row.get("selected_option") == target for row in control_rows)
            control_rate = control_target_count / control_n * 100 if control_n else None

            for variant in variants:
                rows = [row for row in segment_rows if row["variant"] == variant]
                valid_n = len(rows)
                counts = Counter(row.get("selected_option") for row in rows)
                target_count = counts.get(target, 0)
                target_rate = target_count / valid_n * 100 if valid_n else None
                effect = None
                if target_rate is not None and control_rate is not None:
                    effect = target_rate - control_rate

                result: dict[str, Any] = {
                    "Demographic": dimension,
                    "Segment": segment,
                    "Variant": variant,
                    "Control": variant == control,
                    "Valid N": valid_n,
                    "Target count": target_count,
                    "Target rate (%)": target_rate,
                    "Effect vs control (pp)": effect,
                }
                for option in options:
                    count = counts.get(option, 0)
                    result[f"{option} count"] = count
                    result[f"{option} rate (%)"] = count / valid_n * 100 if valid_n else None
                output.append(result)
    return output


def profile_balance_rows(
    config: Mapping[str, Any],
    planned_jobs: Sequence[Mapping[str, Any]],
    accepted_rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    variants = [variant["name"] for variant in config["variants"]]
    for dimension in config["demographic_dimensions"]:
        for segment in config["demographic_distributions"][dimension].keys():
            for variant in variants:
                planned = sum(
                    1
                    for job in planned_jobs
                    if job["variant"] == variant
                    and job.get("profile_segments", {}).get(dimension) == segment
                )
                valid = sum(
                    1
                    for row in accepted_rows
                    if row["variant"] == variant
                    and row.get("profile_segments", {}).get(dimension) == segment
                )
                output.append(
                    {
                        "Demographic": dimension,
                        "Segment": segment,
                        "Variant": variant,
                        "Planned N": planned,
                        "Valid N": valid,
                    }
                )
    return output


def _excel_safe(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False)
    if isinstance(value, str):
        if len(value) > 32_760:
            value = value[:32_730] + "\n[TRUNCATED FOR EXCEL CELL LIMIT]"
        if value.startswith(("=", "+", "-", "@")):
            value = "'" + value
    return value


def _format_worksheet(ws, freeze_panes: str | None = None) -> None:
    if freeze_panes:
        ws.freeze_panes = freeze_panes
    ws.sheet_view.showGridLines = False
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 55)


def _write_table(ws, start_row: int, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> int:
    header_fill = PatternFill("solid", fgColor="171B23")
    accent_fill = PatternFill("solid", fgColor="00FF9B")
    header_font = Font(color="F3F4F6", bold=True)
    accent_font = Font(color="000000", bold=True)
    thin_gray = Side(style="thin", color="39404C")

    for column, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, column, header)
        cell.fill = accent_fill if column == 1 else header_fill
        cell.font = accent_font if column == 1 else header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    for row_index, row in enumerate(rows, start=start_row + 1):
        for column, header in enumerate(headers, start=1):
            cell = ws.cell(row_index, column, _excel_safe(row.get(header)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header.endswith("(%)") or header.endswith("(pp)"):
                cell.number_format = "0.0"
            elif header in {"Cost", "Calculated API cost"}:
                cell.number_format = '$0.0000'
    return start_row + 1 + len(rows)


def create_excel_export(
    config: Mapping[str, Any],
    planned_jobs: Sequence[Mapping[str, Any]],
    run_data: Mapping[str, Any],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Setup"

    dark_fill = PatternFill("solid", fgColor="171B23")
    accent_fill = PatternFill("solid", fgColor="00FF9B")
    white_font = Font(color="F3F4F6", bold=True)
    black_bold = Font(color="000000", bold=True)
    section_font = Font(color="000000", bold=True, size=12)

    def section(title: str, row: int) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row, 1, title)
        cell.fill = accent_fill
        cell.font = section_font
        return row + 1

    def key_value(label: str, value: Any, row: int) -> int:
        label_cell = ws.cell(row, 1, label)
        label_cell.fill = dark_fill
        label_cell.font = white_font
        value_cell = ws.cell(row, 2, _excel_safe(value))
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        return row + 1

    row = 1
    row = section("Run metadata", row)
    row = key_value("Experiment name", config["experiment_name"], row)
    row = key_value("Run timestamp (UTC)", run_data.get("run_timestamp"), row)
    row = key_value("Model", MODEL_ID, row)
    row = key_value("Plan seed", config.get("plan_seed"), row)
    row = key_value("Total target valid N", config["total_simulations"], row)
    row = key_value("Control variant", config["control_variant"], row)
    row = key_value("Representative U.S. preset", config["use_us_preset"], row)
    row = key_value("Preset note", US_PRESET_NOTE if config["use_us_preset"] else "Custom marginal distributions", row)
    row += 1

    row = section("Experiment context", row)
    row = key_value("Line of business", config["line_of_business"], row)
    row = key_value("Customer journey type", config["customer_journey_type"], row)
    row = key_value("Additional context", config["context"], row)
    row += 1

    row = section("Outcome", row)
    row = key_value("Outcome question", config["outcome_question"], row)
    row = key_value("Answer options", "\n".join(config["answer_options"]), row)
    row = key_value("Target behavior", config["target_behavior"], row)
    row += 1

    row = section("Model and sampling settings", row)
    row = key_value("Temperature", config["settings"]["temperature"], row)
    row = key_value("Top-p", config["settings"]["top_p"], row)
    row = key_value("Top-k", config["settings"]["top_k"] if config["settings"]["use_top_k"] else "Disabled", row)
    row = key_value("Max generated tokens", MAX_NEW_TOKENS, row)
    row = key_value("Max slot attempts", MAX_SLOT_ATTEMPTS, row)
    row = key_value("Variant assignment", "Exact requested allocation; jobs randomly interleaved", row)
    row = key_value("Demographic balancing", "Marginal quotas reproduced separately within each variant", row)
    row = key_value("Response-order control", "Cyclic Latin counterbalancing of option-to-code mappings", row)
    row += 1

    row = section("Variants", row)
    variant_headers = ["Variant", "Allocation (%)", "Target N", "Control", "Content"]
    variant_rows = [
        {
            "Variant": variant["name"],
            "Allocation (%)": variant["allocation"],
            "Target N": variant["count"],
            "Control": variant["name"] == config["control_variant"],
            "Content": variant["text"],
        }
        for variant in config["variants"]
    ]
    row = _write_table(ws, row, variant_headers, variant_rows) + 1

    row = section("Demographic marginal distributions", row)
    demographic_rows = []
    for dimension in config["demographic_dimensions"]:
        for category, percentage in config["demographic_distributions"][dimension].items():
            demographic_rows.append(
                {"Demographic": dimension, "Category": category, "Percentage": percentage}
            )
    row = _write_table(ws, row, ["Demographic", "Category", "Percentage"], demographic_rows) + 1

    row = section("Run quality and usage", row)
    quality_pairs = [
        ("Valid responses", run_data.get("valid_total")),
        ("Target responses", run_data.get("total_target")),
        ("Invalid model outputs", run_data.get("invalid_outputs")),
        ("API-failed slot attempts", run_data.get("api_failed_attempts")),
        ("Temporary API retry events", run_data.get("retry_events")),
        ("Simulation attempts", run_data.get("simulation_attempts")),
        ("Underlying HTTP requests", run_data.get("http_requests")),
        ("Input tokens", run_data.get("input_tokens")),
        ("Output tokens", run_data.get("output_tokens")),
        ("Calculated API cost", run_data.get("actual_cost")),
    ]
    for label, value in quality_pairs:
        row = key_value(label, value, row)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 90
    ws.sheet_view.showGridLines = False

    overall = wb.create_sheet("Overall Results")
    overall_rows = overall_result_rows(config, run_data["accepted_rows"])
    overall_headers = list(overall_rows[0].keys()) if overall_rows else ["Variant"]
    _write_table(overall, 1, overall_headers, overall_rows)
    _format_worksheet(overall, "A2")
    overall.auto_filter.ref = overall.dimensions

    segments = wb.create_sheet("Segment Results")
    segment_rows = segment_result_rows(config, run_data["accepted_rows"])
    segment_headers = list(segment_rows[0].keys()) if segment_rows else ["Demographic"]
    _write_table(segments, 1, segment_headers, segment_rows)
    _format_worksheet(segments, "A2")
    segments.auto_filter.ref = segments.dimensions

    balance = wb.create_sheet("Profile Balance")
    balance_rows = profile_balance_rows(config, planned_jobs, run_data["accepted_rows"])
    balance_headers = list(balance_rows[0].keys()) if balance_rows else ["Demographic"]
    _write_table(balance, 1, balance_headers, balance_rows)
    _format_worksheet(balance, "A2")
    balance.auto_filter.ref = balance.dimensions

    raw = wb.create_sheet("Raw Simulations")
    raw_rows: list[dict[str, Any]] = []
    for result in run_data["rows"]:
        row_data: dict[str, Any] = {
            "Attempt ID": result.get("attempt_id"),
            "Slot ID": result.get("slot_id"),
            "Respondent ID": result.get("respondent_id"),
            "Slot attempt": result.get("slot_attempt"),
            "Variant": result.get("variant"),
            "Status": result.get("status"),
            "Selected code": result.get("selected_code"),
            "Selected option": result.get("selected_option"),
            "Raw completion": result.get("raw_completion"),
            "Error": result.get("error"),
            "Option order": result.get("option_order"),
            "Code-to-option mapping": result.get("code_to_option"),
            "Prompt tokens": result.get("prompt_tokens"),
            "Completion tokens": result.get("completion_tokens"),
            "Cost": result.get("cost"),
            "HTTP attempts": result.get("http_attempts"),
            "Temporary retries": result.get("retry_events"),
            "Finish reason": result.get("finish_reason"),
        }
        for dimension in config["demographic_dimensions"]:
            row_data[f"{dimension} segment"] = result.get("profile_segments", {}).get(dimension)
            row_data[f"{dimension} prompt value"] = result.get("profile_values", {}).get(dimension)
        raw_rows.append(row_data)

    raw_headers = list(raw_rows[0].keys()) if raw_rows else ["Attempt ID"]
    _write_table(raw, 1, raw_headers, raw_rows)
    _format_worksheet(raw, "A2")
    raw.auto_filter.ref = raw.dimensions

    sources = wb.create_sheet("Sources and Notes")
    source_rows = [
        {"Type": "Model card", "Source": "https://huggingface.co/socratesft/socrates-qwen2.5-14b-sft"},
        {"Type": "Socrates paper", "Source": "https://arxiv.org/abs/2509.05830"},
        {"Type": "Preset note", "Source": US_PRESET_NOTE},
    ] + [{"Type": "U.S. preset source", "Source": source} for source in US_PRESET_SOURCES]
    _write_table(sources, 1, ["Type", "Source"], source_rows)
    _format_worksheet(sources, "A2")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_run_timestamp() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()
